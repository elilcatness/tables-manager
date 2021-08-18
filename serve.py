import csv
import os

import openpyxl

EXTENSIONS = ['csv', 'xls', 'xlsx']


def contains(x, y):
    if isinstance(x, str):
        x = x.lower()
    if isinstance(y, str):
        y = y.lower()
    return y in x


OPERATORS = {'>': (lambda x, y: x > y, 'Больше, чем'),
             '<': (lambda x, y: x < y, 'Меньше, чем'),
             '=': (lambda x, y: x == y, 'Равно'),
             'contains': (contains, 'Содержит'),
             '~contains': (lambda x, y: not contains(x, y), 'Не содержит')}

DEFAULT_DELIMITER = ';'


class ExcelIterator:
    def __init__(self, excel):
        self.workbook = excel.get_workbook()
        self.current_row = excel.get_start()

    def __next__(self):
        sheet = self.workbook.active
        if self.current_row <= sheet.max_row:
            row = [sheet.cell(self.current_row, col).value for col in range(1, sheet.max_column + 1)]
            self.current_row += 1
            return row
        raise StopIteration


class ExcelReader:
    def __init__(self, filename, start_from=1):
        self.workbook = openpyxl.open(filename)
        self.start_from = start_from

    def get_workbook(self):
        return self.workbook

    def get_start(self):
        return self.start_from

    def __iter__(self):
        return ExcelIterator(self)


def get_headers(filename, delimiter=';'):
    extension = get_extension(filename)
    if extension not in EXTENSIONS:
        return None
    if extension == 'csv':
        f = open(filename, newline='', encoding='utf-8')
        data_reader = csv.reader(f, delimiter=delimiter)
    else:
        f = None
        data_reader = iter(ExcelReader(filename))
    headers = next(data_reader)
    if f:
        f.close()
    return headers


def process_query(statements, indexes=False):
    logical_indexes = [i for i in range(len(statements)) if statements[i] in ('and', 'or')]
    processed_statements = []
    for idx in logical_indexes:
        if idx < len(statements) - 1:
            statements.insert(idx + 1, statements[0])
        else:
            statements.append(statements[0])
        processed_statements.extend([tuple(statements[idx - 3:idx]), statements[idx],
                                     tuple(statements[idx + 1:idx + 4])])
    if not logical_indexes:
        processed_statements.append(tuple(statements))
    if indexes:
        for i in range(len(processed_statements)):
            if isinstance(processed_statements[i], tuple):
                processed_statements[i] = (int(processed_statements[i][0]),) + processed_statements[i][1:]
    return processed_statements


def parse_query(query):
    statements = []
    idx = -1
    statement = ''
    while idx < len(query) - 1:
        idx += 1
        if query[idx] == '"':
            idx += 1
            if idx == len(query):
                break
            statement += query[idx]
            while idx < len(query) - 1 and query[idx + 1] != '"':
                idx += 1
                statement += query[idx]
            statements.append(statement)
            statement = ''
            idx = idx + 1 if idx < len(query) - 1 else idx
        elif query[idx] == ' ':
            if statement:
                statements.append(statement)
                statement = ''
        else:
            statement += query[idx]
    if statement:
        statements.append(statement)
    return statements


def show_help():
    print('\n'.join([f'{key}: {val}' for key, val in OPERATORS.items()]))


def get_extension(filename):
    return filename.split('.')[-1]


def is_empty(filename):
    try:
        if get_extension(filename) == 'csv':
            with open(filename, encoding='utf-8') as f:
                return not bool(f.read())
        else:
            try:
                next(iter(ExcelReader(filename)))
            except StopIteration:
                return True
            return False
    except UnicodeError:
        return True


def get_reader(filename, delimiter=';'):
    extension = get_extension(filename)
    if get_extension(filename) not in EXTENSIONS:
        return None
    if extension == 'csv':
        file = open(filename, newline='', encoding='utf-8')
        reader = csv.reader(file, delimiter=delimiter)
        return reader, file
    return ExcelReader(filename), None


def validate_by_filters(data, filters):
    query = ''
    for i, value in enumerate(filters.values()):
        for statement in value:
            if isinstance(statement, str):
                query += f' {statement} '
            else:
                key, operator, val = statement
                try:
                    float(val)
                except ValueError:
                    val = f'"{val}"'
                subquery = f'OPERATORS["{operator}"][0](data[{key}], {val})'
                query += subquery
        if i < len(filters.values()) - 1:
            query += ' and '
    return eval(query)


def split_files(filenames, rows_count, data, add_headers=True):
    for j, filename in enumerate(filenames):
        print(f'Разбиваем {filename}')
        headers, filters, delimiter = (data[filename][key] for key in ('headers', 'filters', 'delimiter'))
        count, file_count = 0, 0
        reader, main_file = get_reader(filename, delimiter)
        last_message = ''
        skip = False
        csv_file, writer = None, None
        for i, row in enumerate(reader):
            if count % rows_count == 0 and not skip:
                if file_count > 0:
                    last_message = f'Записан {".".join(filename.split(".")[:-1])}_{file_count}.csv'
                    print(last_message)
                file_count += 1
                if csv_file:
                    csv_file.close()
                csv_file = open(f'{".".join(filename.split(".")[:-1])}_{file_count}.csv',
                                'w', newline='', encoding='utf-8')
                writer = csv.writer(csv_file, delimiter=delimiter)
                if headers and add_headers:
                    writer.writerow(headers)
            if filters:
                row_data = {key: val for key, val
                            in zip(headers if headers
                                   else [i for i in range(len(row))], row)}
                skip = True
                if not validate_by_filters(row_data, filters):
                    continue
            if row != headers:
                writer.writerow(row)
                count += 1
                skip = False
            else:
                skip = True
        new_msg = f'Записан {".".join(filename.split(".")[:-1])}_{file_count}.csv'
        if new_msg != last_message:
            print(new_msg)
        if main_file:
            main_file.close()
        if j < len(filenames) - 1:
            print('\n' + '#' * 50 + '\n')


def get_correct_filenames():
    correct_filenames = [f_name for f_name in os.listdir() if get_extension(f_name) in EXTENSIONS]
    correct_filenames = list(filter(lambda f_name: not is_empty(f_name), correct_filenames))
    return correct_filenames


def manage_split_files(correct_filenames):
    add_headers = input('Добавлять ли в каждый выходной файл шапку с заголовками? (y\\n) '
                        ).lower() == 'y'
    while True:
        try:
            rows_count = int(input('Введите количество строк в каждом файле: '))
            break
        except ValueError:
            print('Неверный формат, введите число ещё раз')
    data = {filename: {'headers': [], 'filters': {}, 'delimiter': ';'} for filename in correct_filenames}
    for filename in correct_filenames:
        file_has_headers = input(f'Имеет ли {filename} заголовки? (y\\n) ').lower() == 'y'
        if file_has_headers:
            data[filename]['headers'] = get_headers(filename, delimiter=';')
        delimiter = input(f'Введите разделитель в файле {filename} '
                          f'(нажмите Enter для значения по умолчанию: "{DEFAULT_DELIMITER}"): ')
        data[filename]['delimiter'] = delimiter if delimiter else DEFAULT_DELIMITER
    set_filters = input('Будете ли Вы устанавливать фильтры? (y\\n) ').lower() == 'y'
    while set_filters:
        print(f"{'*' * 20} Меню фильтрации {'*' * 20}")
        print(' '.join(correct_filenames))
        exit_menu = False
        while True:
            filename = input('Введите название файла (для выхода нажмите Enter): ')
            if not filename:
                exit_menu = True
                break
            if filename.strip() in correct_filenames:
                break
            else:
                print('Неверное название файла')
        if exit_menu:
            break
        headers = data[filename]['headers']
        if headers:
            print(f"Столбцы: {';'.join(headers)}")
        else:
            print(' '.join([str(i) for i in
                            range(1, len(get_headers(
                                filename, delimiter=data[filename]['delimiter'])) + 1)]))
        print('Введите запросы для фильтрации в виде: НАЗВАНИЕ_СТОЛБЦА ОПЕРАТОР ЗНАЧЕНИЕ '
              '(and/or ОПЕРАТОР ЗНАЧЕНЕИЕ)n',
              f'Для выхода из меню фильтрации для файла {filename} нажмите Enter',
              'Для получения операторов и их функций введите /help', sep='\n', end='\n\n')
        while True:
            filter_query = input()
            if not filter_query:
                break
            if filter_query == '/help':
                show_help()
                continue
            try:
                statements = parse_query(filter_query)
                operators_statements = statements[1::3]
                if not statements:
                    print('Значение фильтра не может быть пустым')
                elif headers and statements[0].lower() not in map(str.lower, headers):
                    print('Указан несуществующий столбец')
                elif not headers and not statements[0].isdigit():
                    print('Номер столбца должен быть целочисленным')
                elif len(statements) < 3:
                    print('Недостаточно аргументов')
                elif len(statements) > 3 and ('and' not in statements and 'or' not in statements):
                    print('Упущены логические И/ИЛИ')
                elif any(map(lambda statement: statement not in OPERATORS.keys(), operators_statements)):
                    print(', '.join([f"{statement}" for statement in operators_statements]),
                          'Are not operators', sep=' - ')
                else:
                    if headers:
                        data[filename]['filters'][
                            headers[headers.index(statements[0])]] = process_query(statements)
                    else:
                        statements[0] = int(statements[0]) - 1
                        data[filename]['filters'][statements[0]] = process_query(statements,
                                                                                 indexes=True)
                    continue
            except ValueError:
                print('Неверно введены значения')
    print()
    split_files(correct_filenames, rows_count, data, add_headers)


def unite_files(output_filename, headers_filename, filenames, delimiter=';', no_headers=False):
    if not no_headers:
        skip = not bool(headers_filename)
        headers = get_headers(headers_filename if headers_filename else filenames[0], delimiter)
    else:
        skip, headers = False, []
    with open(output_filename, 'w', newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, delimiter=delimiter)
        if headers:
            writer.writerow(headers)
        for filename in filenames:
            reader, main_file = get_reader(filename, delimiter)
            file_headers = []
            for i, row in enumerate(reader):
                if not no_headers and (skip or filename == headers_filename):
                    if i == 0:
                        file_headers = row
                        continue
                    data = {key: val for key, val in zip(file_headers, row) if key in headers}
                    row = [data.get(key) for key in headers]
                row = row if len(row) <= len(headers) or no_headers else row[:len(headers)]
                writer.writerow(row)
            print(f'Записан {filename} в {output_filename}')
            if main_file:
                main_file.close()


def manage_unite_files(correct_filenames):
    headers_filename = ''
    no_headers = False
    if input('В каждом ли файле первая строка – заголовки шапки? (y\\n) ').lower() != 'y':
        print(' '.join(correct_filenames))
        while True:
            headers_filename = input('Введите название файла, из которого брать заголовки '
                                     '(если ни в одном нет - нажмите Enter): ')
            if not headers_filename:
                no_headers = True
                break
            elif headers_filename not in correct_filenames:
                print('Файла нет в директории')
            else:
                break
    delimiter = input('Введите разделитель во всех файлах '
                      f'(нажмите Enter для значения по умолчанию: "{DEFAULT_DELIMITER}"): ')
    output_filename = input('Введите название выходного файла: ')
    if get_extension(output_filename) != 'csv':
        return print('Выходной файл должен иметь расширение csv')
    unite_files(output_filename, headers_filename,
                correct_filenames, delimiter if delimiter else DEFAULT_DELIMITER,
                no_headers)


def main():
    correct_filenames = get_correct_filenames()
    if not correct_filenames:
        return f'В директории нет файлов с подходящим расширением ({", ".join(EXTENSIONS)})'
    actions = {'разбить': manage_split_files, '1': manage_split_files,
               'объединить': manage_unite_files, '2': manage_unite_files}
    while True:
        action = input('Выберите действие: разбить(1) / объединить(2): ')
        if actions.get(action.lower()):
            action = actions[action]
            break
        else:
            print('Неверное значение', end='\n\n')
    action(correct_filenames)


if __name__ == '__main__':
    callback = main()
    if callback:
        print(callback)
